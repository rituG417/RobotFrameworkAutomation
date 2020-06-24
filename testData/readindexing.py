#install package File --> Setting --> project--> pythoninterpreter -> openpyxl-- install
import openpyxl
#loading workbook
wb=openpyxl.load_workbook("C:\\Users\\ritu.garg\\PycharmProjects\\AutomationDay1\\testData\\TestData.xlsx")

#defining function to return max no. of ros in a excel sheet
def fetch_numberofrows(sheetName):
    sh = wb[sheetName]
    return sh.max_row
#max_row: predefined method tells the max no. of rows in a sheet.

def fetch_cellData(sheetName,row,cell):
    sh = wb[sheetName]
    #int used to convert string type into int from excel.
    cell = sh.cell(int(row),int(cell))
    return cell.value

#x=fetch_cellData('ValidLogin',1,1)
#print(x)

